#!/usr/bin/env python3
"""Stream-convert Online Retail II Excel workbook to CSV.

Pandas can be slow for very large Excel files. This script uses openpyxl in read_only
mode to stream rows and write a single CSV combining all sheets.

Usage:
  python etl/convert_xlsx_to_csv.py --input data/raw/online_retail_II.xlsx --output data/raw/online_retail_II_combined.csv
"""

import argparse
import csv
from openpyxl import load_workbook


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True)
    ap.add_argument("--output", required=True)
    args = ap.parse_args()

    wb = load_workbook(args.input, read_only=True, data_only=True)
    sheet_names = wb.sheetnames

    with open(args.output, "w", newline="", encoding="utf-8") as f:
        writer = None
        for sname in sheet_names:
            ws = wb[sname]
            rows = ws.iter_rows(values_only=True)
            header = next(rows)
            if writer is None:
                writer = csv.writer(f)
                writer.writerow(header)
            # If later sheets have identical header, skip header row already written
            for r in rows:
                writer.writerow(r)


if __name__ == "__main__":
    main()
